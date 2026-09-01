import base64
import io
import os
import re
import shutil
import subprocess
import tempfile
import unicodedata
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

MARKERS = {
    "ONE POINT LESSON": 6,
    "OPL N": 5,
    "TITOLO": 3,
    "REPARTO": 3,
    "LINEA": 3,
    "PROBLEMA": 2,
    "MIGLIORAMENTO": 2,
    "CAUSA": 1,
}

LABELS = {
    "numero": ["OPL N", "OPL NO", "N OPL", "NUMERO OPL"],
    "titolo": ["TITOLO"],
    "tipo_opl": ["TIPO"],
    "autore": ["AUTORE", "COMPILATO DA"],
    "reparto_originale": ["REPARTO"],
    "linea_originale": ["LINEA"],
    "problema": ["PROBLEMA"],
    "causa": ["CAUSA"],
    "miglioramento": ["MIGLIORAMENTO"],
    "data_documento": ["DATA"],
    "area_opl": ["AREA OPL COLORE BORDO", "AREA OPL"],
}

EMPTY_VALUES = {"", "-", "--", "---", "- - -", "N/A", "NONE"}


def normalize_text(value):
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    text = str(value).replace("\n", " ").replace("\r", " ").strip()
    return re.sub(r"\s+", " ", text)


def normalize_key(value):
    text = normalize_text(value).upper()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"[^A-Z0-9]+", " ", text).strip()


def is_real_value(value):
    return normalize_key(value) not in EMPTY_VALUES


def sheet_cells(ws):
    cells = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value not in (None, ""):
                cells.append(cell)
    return cells


def sheet_score(ws):
    values = [normalize_key(cell.value) for cell in sheet_cells(ws)]
    score = 0
    for marker, weight in MARKERS.items():
        if any(marker in value for value in values):
            score += weight
    core = 0
    for label in ("OPL N", "TITOLO", "REPARTO", "LINEA"):
        if any(label in value for value in values):
            core += 1
    real_values = sum(1 for value in values if value and value not in EMPTY_VALUES)
    score += min(real_values // 10, 4)
    if ws.sheet_state != "visible":
        score -= 8
    if core < 2:
        score -= 10
    return score


def select_opl_sheet(workbook):
    ranked = sorted(
        ((sheet_score(ws), index, ws) for index, ws in enumerate(workbook.worksheets)),
        key=lambda item: (item[0], -item[1]),
        reverse=True,
    )
    if not ranked or ranked[0][0] < 5:
        raise ValueError("Nessun foglio OPL riconoscibile")
    return ranked[0][2], [{"sheet": item[2].title, "score": item[0]} for item in ranked]


def merged_anchor_value(ws, row, col):
    value = ws.cell(row=row, column=col).value
    if value not in (None, ""):
        return value
    coordinate = ws.cell(row=row, column=col).coordinate
    for merged_range in ws.merged_cells.ranges:
        if coordinate in merged_range:
            return ws.cell(merged_range.min_row, merged_range.min_col).value
    return None


def find_value_near_label(ws, aliases):
    aliases = [normalize_key(alias) for alias in aliases]
    for cell in sheet_cells(ws):
        key = normalize_key(cell.value)
        if not any(alias == key or key.startswith(alias + " ") for alias in aliases):
            continue
        inline = normalize_text(cell.value)
        if ":" in inline:
            suffix = inline.split(":", 1)[1].strip()
            if is_real_value(suffix):
                return suffix
        candidates = []
        for offset in range(1, 7):
            candidates.append(merged_anchor_value(ws, cell.row, cell.column + offset))
        for offset in range(1, 4):
            candidates.append(merged_anchor_value(ws, cell.row + offset, cell.column))
        for candidate in candidates:
            if is_real_value(candidate):
                candidate_key = normalize_key(candidate)
                if not any(candidate_key == alias for alias in aliases):
                    return normalize_text(candidate)
    return ""



def find_date_value(ws):
    candidates = []
    for cell in sheet_cells(ws):
        key = normalize_key(cell.value)
        if key != "DATA" and not key.startswith("DATA "):
            continue
        inline = normalize_text(cell.value)
        match = re.search(r"(\d{1,2}/\d{1,2}/\d{4}|\d{4}-\d{1,2}-\d{1,2})", inline)
        if match:
            candidates.append(match.group(1))
        for offset in range(1, 7):
            candidates.append(merged_anchor_value(ws, cell.row, cell.column + offset))
        for offset in range(1, 4):
            candidates.append(merged_anchor_value(ws, cell.row + offset, cell.column))
    for candidate in candidates:
        if isinstance(candidate, (datetime, date)):
            return candidate
        text = normalize_text(candidate)
        if re.fullmatch(r"\d{1,2}/\d{1,2}/\d{4}", text) or re.fullmatch(r"\d{4}-\d{1,2}-\d{1,2}", text):
            return candidate
    return ""

def number_from_filename(filename):
    match = re.search(r"(?i)\bOPL\s*[-_ ]*([0-9]{1,6})(?:\s*[_/-]\s*([0-9]{1,3}))?", filename)
    if not match:
        return ""
    number = f"OPL-{match.group(1)}"
    if match.group(2):
        number += f"_{match.group(2)}"
    return number


def normalize_number(value, filename):
    text = normalize_text(value).upper().replace(" ", "")
    text = text.replace("OPLN°", "OPL-").replace("OPLNO.", "OPL-")
    if not text:
        return number_from_filename(filename)
    match = re.search(r"(?:OPL[-_]*)?([0-9]{1,6})(?:[_/-]([0-9]{1,3}))?", text)
    if not match:
        return number_from_filename(filename)
    number = f"OPL-{match.group(1)}"
    if match.group(2):
        number += f"_{match.group(2)}"
    return number


def normalize_department(value):
    original = normalize_text(value)
    key = normalize_key(original)
    aliases = {
        "MOD NORD": "Modellaggio Nord",
        "MODELLAGGIO NORD": "Modellaggio Nord",
        "MOD SUD": "Modellaggio Sud",
        "MODELLAGGIO SUD": "Modellaggio Sud",
        "CONFEZIONE": "Confezione",
    }
    return aliases.get(key, original.title() if original else "")


def normalize_line(value):
    original = normalize_text(value)
    key = normalize_key(original)
    match = re.search(r"(BINDLER|BETTI)\s*([0-9]+)", key)
    if match:
        return f"{match.group(1).title()} {match.group(2)}"
    return original.title() if original else ""


def parse_date(value):
    if isinstance(value, datetime):
        return value.date().isoformat(), False
    if isinstance(value, date):
        return value.isoformat(), False
    text = normalize_text(value)
    if not text:
        return "", False
    for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date().isoformat(), False
        except ValueError:
            pass
    try:
        parsed = datetime.strptime(text, "%m/%d/%Y").date()
        ambiguous = parsed.day <= 12 and parsed.month <= 12
        return parsed.isoformat(), ambiguous
    except ValueError:
        return text, True


def render_selected_sheet(contents, filename, selected_sheet):
    if not shutil.which("libreoffice") or not shutil.which("pdftoppm"):
        return None, "Anteprima non disponibile sul server corrente"
    with tempfile.TemporaryDirectory() as temp_dir:
        temp = Path(temp_dir)
        source = temp / Path(filename).name
        source.write_bytes(contents)
        workbook = load_workbook(source)
        for ws in workbook.worksheets:
            ws.sheet_state = "visible" if ws.title == selected_sheet else "hidden"
        workbook.active = workbook.sheetnames.index(selected_sheet)
        prepared = temp / f"prepared_{source.name}"
        workbook.save(prepared)
        result = subprocess.run(
            ["libreoffice", "--headless", "--convert-to", "pdf", "--outdir", str(temp), str(prepared)],
            capture_output=True,
            text=True,
            timeout=90,
        )
        pdf = temp / f"{prepared.stem}.pdf"
        if result.returncode != 0 or not pdf.exists():
            return None, "Conversione anteprima non riuscita"
        png_base = temp / "preview"
        result = subprocess.run(
            ["pdftoppm", "-f", "1", "-singlefile", "-png", "-r", "120", str(pdf), str(png_base)],
            capture_output=True,
            text=True,
            timeout=90,
        )
        png = temp / "preview.png"
        if result.returncode != 0 or not png.exists():
            return None, "Creazione immagine non riuscita"
        encoded = base64.b64encode(png.read_bytes()).decode("ascii")
        return f"data:image/png;base64,{encoded}", None


def analyze_excel(contents, filename, include_preview=True):
    workbook = load_workbook(io.BytesIO(contents), data_only=True, read_only=False)
    worksheet, ranking = select_opl_sheet(workbook)
    extracted = {field: find_value_near_label(worksheet, aliases) for field, aliases in LABELS.items()}
    extracted["data_documento"] = find_date_value(worksheet)
    extracted["numero"] = normalize_number(extracted["numero"], filename)
    extracted["reparto"] = normalize_department(extracted.pop("reparto_originale"))
    extracted["linea"] = normalize_line(extracted.pop("linea_originale"))
    parsed_date, ambiguous_date = parse_date(extracted["data_documento"])
    extracted["data_documento"] = parsed_date

    warnings = []
    required = ["numero", "titolo", "reparto", "linea"]
    missing = [field for field in required if not extracted.get(field)]
    if missing:
        warnings.append("Campi da verificare: " + ", ".join(missing))
    if ambiguous_date:
        warnings.append("Formato data ambiguo o non riconosciuto")
    if extracted["numero"] and extracted["numero"] not in filename.upper().replace(" ", ""):
        filename_number = number_from_filename(filename)
        if filename_number and filename_number != extracted["numero"]:
            warnings.append(f"Numero nel file diverso dal nome: {filename_number}")

    found = sum(1 for field in required if extracted.get(field))
    optional = ["autore", "tipo_opl", "problema", "causa", "miglioramento", "data_documento"]
    found_optional = sum(1 for field in optional if extracted.get(field))
    confidence = round(min(100, 45 + found * 10 + found_optional * 2.5 - len(warnings) * 5), 1)

    preview = None
    preview_warning = None
    if include_preview:
        preview, preview_warning = render_selected_sheet(contents, filename, worksheet.title)
        if preview_warning:
            warnings.append(preview_warning)

    return {
        "filename": filename,
        "sheet": worksheet.title,
        "sheet_ranking": ranking,
        "confidence": confidence,
        "warnings": warnings,
        "preview": preview,
        **extracted,
    }
